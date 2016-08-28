from openpyxl import load_workbook
import numpy as np

from openpyxl import load_workbook
wb = load_workbook(filename='data/CleanDataKeyEntry.xlsx', read_only=True, data_only=True)

ws = wb.get_sheet_by_name('MainData')
A = np.array([[i.value for i in j] for j in ws.rows])

columns = A[0,:]
story_text_column_idx = np.argwhere(columns == 'Story_Text')[0]
# remove header row
data = np.array(A[range(1,A.shape[0]),:])
data = data[range(823),:]
documents = [data.item((i,story_text_column_idx)) for i in range(data.shape[0])]
